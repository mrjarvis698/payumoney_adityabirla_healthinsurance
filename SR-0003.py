import json, shutil, warnings, time
import os
from os import path
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, NoSuchFrameException, NoSuchWindowException, UnexpectedAlertPresentException, WebDriverException, NoSuchWindowException, ElementClickInterceptedException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys
import getpass

def clearConsole():
    command = 'clear'
    if os.name in ('nt', 'dos'):  # If Machine is running on Windows, use cls
        command = 'cls'
    os.system(command)
os.system('mode con: cols=65 lines=15')

appdata_path = os.getenv('APPDATA')
# License
licence_pass = "SR1002"
licence_dir = path.exists(appdata_path + "/health/")
global license_verify
if licence_dir == True :
    licence_sheet = path.exists(appdata_path + "/health/license.json")
    if licence_sheet == True:
        license_file_path = appdata_path + "/health/license.json"
        json_file = open(license_file_path)
        data = json.load(json_file)
        licence_check = data ['license']
        if licence_check == licence_pass:
            print("Licence Verifed")
            license_verify = True
        else :
            shutil.rmtree(appdata_path + "/health", ignore_errors=True)
            license = getpass.getpass()
            if license == licence_pass:
                cache_path = os.path.join(str(os.getcwd()), appdata_path + "/health")
                os.mkdir(cache_path)
                dictionary = {"license" : license}
                json_object = json.dumps(dictionary, indent = 1)
                with open(appdata_path + "/health/license.json", "w") as outfile:
                    outfile.write(json_object)
                license_verify = True
            else:
                license_verify = False
                clearConsole()
                print("Invalid Licence")
                time.sleep(5)
    else:
        license = getpass.getpass()
        if license == licence_pass:
            dictionary = {"license" : license}
            json_object = json.dumps(dictionary, indent = 1)
            with open(appdata_path + "/health/license.json", "w") as outfile:
                outfile.write(json_object)
            license_verify = True
        else:
            license_verify = False
            clearConsole()
            print("Invalid Licence")
            time.sleep(5)
else :
    license = getpass.getpass()
    if license == licence_pass:
        cache_path = os.path.join(str(os.getcwd()), appdata_path + "/health")
        os.mkdir(cache_path)
        dictionary = {"license" : license}
        json_object = json.dumps(dictionary, indent = 1)
        with open(appdata_path + "/health/license.json", "w") as outfile:
            outfile.write(json_object)
        license_verify = True
    else:
        license_verify = False
        clearConsole()
        print("Invalid Licence")
        time.sleep(5)

if license_verify == True:
    # Open xlsx file
    open_sheet = path.exists("zaggle_cache/opened_sheet.json")
    global xlsx_file_path
    if open_sheet == True :
        opened_sheet_file_path = "zaggle_cache/opened_sheet.json"
        json_file = open(opened_sheet_file_path)
        data = json.load(json_file)
        xlsx_sheet_check = path.exists(data ['xlsx_file_path'])
        if xlsx_sheet_check == True :
            xlsx_file_path = data ['xlsx_file_path']
        else :
            shutil.rmtree('zaggle_cache', ignore_errors=True)
            xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
            cache_path = os.path.join(str(os.getcwd()), "zaggle_cache")
            dictionary = {"xlsx_file_path" : xlsx_file_path}
            json_object = json.dumps(dictionary, indent = 1)
            with open("zaggle_cache/opened_sheet.json", "w") as outfile:
                outfile.write(json_object)
    else :
        xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
        cache_path = os.path.join(str(os.getcwd()), "zaggle_cache")
        os.mkdir(cache_path)
        dictionary = {"xlsx_file_path" : xlsx_file_path}
        json_object = json.dumps(dictionary, indent = 1)
        with open("zaggle_cache/opened_sheet.json", "w") as outfile:
            outfile.write(json_object)

    # read imported xlsx file path using pandas
    input_workbook = pd.read_excel(xlsx_file_path, sheet_name = 'Sheet1', dtype=str)
    total_input_rows, total_input_cols = input_workbook.shape

    input_col = list(input_workbook.columns.values.tolist())

    input_xlsx_col_A = input_workbook[input_col[0]].values.tolist()
    input_xlsx_col_B = input_workbook[input_col[1]].values.tolist()
    input_xlsx_col_C = input_workbook[input_col[2]].values.tolist()
    input_xlsx_col_D = input_workbook[input_col[3]].values.tolist()
    input_xlsx_col_E = input_workbook[input_col[4]].values.tolist()
    input_xlsx_col_F = input_workbook[input_col[5]].values.tolist()
    input_xlsx_col_G = input_workbook[input_col[6]].values.tolist()
    input_xlsx_col_H = input_workbook[input_col[7]].values.tolist()
    input_xlsx_col_I = input_workbook[input_col[8]].values.tolist()
    input_xlsx_col_J = input_workbook[input_col[9]].values.tolist()

    # get-output sheet to append output
    output_sheet = path.exists("Output.xlsx")
    if output_sheet == True :
        output_sheet_file_path = "Output.xlsx"
    else :
        input_col.append('No.of Transactions')
        input_col.append('Transaction Status')
        input_col.append('Transaction ID')
        output_headers = input_col
        overall_output = Workbook()
        page = overall_output.active
        page.append(output_headers)
        overall_output.save(filename = 'Output.xlsx')
        output_sheet_file_path = "Output.xlsx"

    def cal():
        global output_cc_number
        global done_transactions_wb
        global h
        
        output_load_wb = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', dtype=str)
        output_col = list(output_load_wb.columns.values.tolist())
        output_cc_number = output_load_wb[output_col[4]].values.tolist()
        done_transactions_wb = output_load_wb[output_col[10]].values.tolist()
        total_output_rows, total_output_cols = output_load_wb.shape
        h = total_output_rows - 1
        print ("-"*63,"\nLast Txn Card No. =",output_cc_number[h],"| Last Card no.of Txns =",done_transactions_wb[h])
        print ("-"*63)

    def output_save():
        global output_wb, entry_list
        entry_list = [[input_xlsx_col_A[i], input_xlsx_col_B[i], input_xlsx_col_C[i], input_xlsx_col_D[i], input_xlsx_col_E[i], input_xlsx_col_F[i], input_xlsx_col_G[i], input_xlsx_col_H[i], input_xlsx_col_I[i], input_xlsx_col_J[i], j + 1, transaction_status, transaction_id]]
        output_wb = load_workbook(output_sheet_file_path)
        page = output_wb.active
        for info in entry_list:
            page.append(info)
        output_wb.save(filename='Output.xlsx')
        clearConsole()
        print ("-"*40,"\nZaggle Cards - Running Card", "| DESK NO. =",input_xlsx_col_J[i])
        print ("-"*40,"\nCard Index =", i+1, "\nCard No =", "XXXX XXXX XXXX", cc_set4, "| Expiry =", input_xlsx_col_H[i], "\nPin =", input_xlsx_col_F[i], "\nStatus =", transaction_status, "\nTransaction no. of this card =", j+1)
        print ("-"*40,"\nCards Done =", i, "| " "Cards Remaining =", total_input_rows - i, "| Total Cards =",total_input_rows)
        print("Elapsed time = " + time.strftime("%H:%M:%S", time.gmtime(time.time() - start_time)))
        print ("-"*40)

    def transactions_continue():
        global card_iteration, transaction_iteration
        try:
            cal()
        except IndexError:
            card_iteration = 0
            transaction_iteration = 0
        else:
            last_txncard =  input_workbook[input_workbook[input_col[4]] == output_cc_number[h]].index[0]
            card_iteration = last_txncard
            transaction_iteration = int(done_transactions_wb[h])

    def cc_number():
        global cc_set1 ,cc_set2, cc_set3, cc_set4
        workbook_cc = input_xlsx_col_E[i]
        cc_set1 = workbook_cc[0:4]
        cc_set2 = workbook_cc[4:8]
        cc_set3 = workbook_cc[8:12]
        cc_set4 = workbook_cc[12:16]

    def cc_expiry():
        global expiry_month, expiry_year, expiry_year1, expiry_year2, expiry_year3, expiry_year4
        workbook_expiry_month = input_xlsx_col_H[i]
        workbook_expiry_year = input_xlsx_col_H[i]
        expiry_month = workbook_expiry_month[:2]
        expiry_year = workbook_expiry_year[3:]
        expiry_year1 = workbook_expiry_year[3]
        expiry_year2 = workbook_expiry_year[4]
        expiry_year3 = workbook_expiry_year[5]
        expiry_year4 = workbook_expiry_year[6]

    def setUp():
        global start_time, driver
        start_time = time.time()
        transactions_continue()
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=500,1080")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-crash-reporter")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-in-process-stack-traces")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_argument("--output=/dev/null")
        executable_path = Service('driver/chromedriver.exe')
        driver = webdriver.Chrome(options=chrome_options, service=executable_path)
        driver.implicitly_wait(3)

    def main():
        global skip, transaction_status, transaction_id, captcha_failed, transaction_failed
        driver.switch_to.window(driver.window_handles[0])
        skip = False
        captcha_failed = False
        transaction_failed = False
        cc_number()
        cc_expiry()
        try:
            driver.get("https://webfront.payu.in/webfront/#/payment/webfrontpage/strawesome-quencher/single/455")
        except WebDriverException:
            print ("No internet - Sleeping For 1 Minute.")
            transaction_status = "No Internet"
            time.sleep(60)
        else:
            driver.get("https://webfront.payu.in/webfront/#/payment/webfrontpage/strawesome-quencher/single/455")
            if skip == False:
                try:
                    driver.find_element(By.NAME, "name")                
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.NAME, "name") .send_keys(input_xlsx_col_A[i])
            else :
                transaction_status = "Null"
                transaction_id = "-"
            
            if skip == False:
                try:
                    driver.find_element(By.NAME, "email")                
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.NAME, "email") .send_keys(input_xlsx_col_D[i])
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.NAME, "phone")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.NAME, "phone").send_keys(input_xlsx_col_C[i])
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.NAME, "Gender")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    Select(driver.find_element(By.NAME, "Gender")).select_by_value('Male')
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.NAME, "Date of Birth")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.NAME, "Date of Birth").send_keys("14/06/2001")
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.XPATH, "//form[@id='payment']/div[2]/div[4]/div[2]/button")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.XPATH, "//form[@id='payment']/div[2]/div[4]/div[2]/button").click()
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                time.sleep(5)
                try:
                    driver.find_element(By.CLASS_NAME, "static-menu")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.CLASS_NAME, "static-menu").click()
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.XPATH, "//*[@id='app']/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/ul/li[4]/span")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.XPATH, "//*[@id='app']/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/ul/li[4]/span").click()
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.ID, "cardNumber")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.ID, "cardNumber").send_keys(cc_set1)
                    driver.find_element(By.ID, "cardNumber").send_keys(cc_set2)
                    driver.find_element(By.ID, "cardNumber").send_keys(cc_set3)
                    driver.find_element(By.ID, "cardNumber").send_keys(cc_set4)
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.ID, "cardExpiry")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.ID, "cardExpiry").send_keys(expiry_month)
                    driver.find_element(By.ID, "cardExpiry").send_keys(expiry_year3)
                    driver.find_element(By.ID, "cardExpiry").send_keys(expiry_year4)
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.ID, "cardCvv")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.ID, "cardCvv").send_keys(input_xlsx_col_G[i])
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.ID, "cardOwnerName")
                except NoSuchElementException:
                    skip = True
                else:
                    skip = False
                    driver.find_element(By.ID, "cardOwnerName").send_keys(input_xlsx_col_A[i])
                    driver.find_element(By.ID, "cardOwnerName").send_keys(Keys.RETURN)
            else :
                transaction_status = "Null"
                transaction_id = "-"

            if skip == False:
                try:
                    driver.find_element(By.ID, "ipin")
                except NoSuchElementException:
                    skip = False
                    try :
                        driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div[2]/div[2]/div/div/span")
                    except NoSuchElementException:
                        skip = True
                        transaction_status = 'Null'
                        transaction_id = "-"
                    else:
                        skip = False
                        transaction_text = driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div[2]/div[2]/div/div/span").text
                        if transaction_text == "Reason of Failure : Payment could not be processed, Try again.":
                            transaction_status = "Failed"
                        transaction_id = "-"
                        transaction_failed = True
                    
                else:
                    skip = False
                    driver.find_element(By.ID, "ipin").send_keys(input_xlsx_col_F[i])
            else :
                transaction_status = 'Null'
                transaction_id = "-"
            
            if skip == False and transaction_failed == False:
                if skip == False:
                    try:
                        driver.find_element(By.ID, "otpbut")
                    except NoSuchElementException:
                        skip = True
                    else:
                        skip = False
                        driver.find_element(By.ID, "otpbut").click()
                else :
                    transaction_status = 'Null'
                    transaction_id = "-"

                if skip == False:
                    try:
                        driver.find_element(By.XPATH, "(.//*[normalize-space(text()) and normalize-space(.)='Aditya Birla Health Insurance Company Limited'])[1]/following::h3[1]")
                    except NoSuchElementException:
                        try:
                            driver.find_element(By.XPATH,"//*[@id='set']/div/div/div[2]/div/div[3]/font")
                        except NoSuchElementException:
                            transaction_status = "Failed"
                            transaction_id = "-"
                        else:
                            skip = False
                            transaction_status = "Incorrect Ipin"
                            transaction_id = "-"
                    else:
                        transaction_status = driver.find_element(By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Aditya Birla Health Insurance Company Limited'])[1]/following::h3[1]").text
                        transaction_id = driver.find_element(By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Your transaction has been successfully done'])[1]/following::i[1]").text
                else :
                    transaction_status = 'Null'
                    transaction_id = "-"


    setUp()
    for i in range (card_iteration , total_input_rows):
        for j in range (transaction_iteration, 4):
            main()
            output_save()
        transaction_iteration = 0
    driver.quit()
